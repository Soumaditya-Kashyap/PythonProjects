import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl

# Function to save data for insertion
def save_data():
    # Get student data from entry fields
    student_name = name_entry.get()
    enrollment_id = id_entry.get()
    dm_score = dm_entry.get()
    daa_score = daa_entry.get()
    os_score = os_entry.get()
    coa_score = coa_entry.get()
    evs_score = evs_entry.get()
    pdp_score = pdp_entry.get()

    # Check if all fields are filled
    if not (student_name and enrollment_id and dm_score and daa_score and os_score and coa_score and evs_score and pdp_score):
        messagebox.showerror("Error", "All fields must be filled.")
        return

    # Save data to Excel file #use path of the excel file according to your local device
    filepath = r"C:\Users\infin\.jupyter\Book1.xlsx"
    
    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        
        # Add headers if the file is empty
        if not sheet["A1"].value:
            sheet["A1"] = "Serial No"
            sheet["B1"] = "Student Name"
            sheet["C1"] = "Enrollment ID"
            sheet["D1"] = "Discrete Mathematics"
            sheet["E1"] = "Design and Analysis of Algorithms"
            sheet["F1"] = "Operating Systems"
            sheet["G1"] = "Computer Organization and Architecture"
            sheet["H1"] = "Environmental Science"
            sheet["I1"] = "Personality Development Program"
        
        # Set column widths
        column_widths = {
            'A': 16,  # Serial No
            'B': 18,  # Student Name
            'C': 18,  # Enrollment ID
            'D': 20,  # DM
            'E': 26,  # DAA
            'F': 20,  # OS
            'G': 26,  # COA
            'H': 18,  # EVS
            'I': 20   # PDP
        }
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
        
        # Find the next empty row
        next_row = 2  # Start from row 2, as row 1 is for headers
        while sheet[f"B{next_row}"].value:
            next_row += 1
        
        # Add student data to the next empty row
        sheet[f"A{next_row}"] = next_row - 1  # Serial No
        sheet[f"B{next_row}"] = student_name
        sheet[f"C{next_row}"] = enrollment_id
        sheet[f"D{next_row}"] = dm_score
        sheet[f"E{next_row}"] = daa_score
        sheet[f"F{next_row}"] = os_score
        sheet[f"G{next_row}"] = coa_score
        sheet[f"H{next_row}"] = evs_score
        sheet[f"I{next_row}"] = pdp_score
        
        # Save the workbook
        wb.save(filepath)
        print("Data saved successfully!")
        
        # Clear entry fields for next entry
        name_entry.delete(0, tk.END)
        id_entry.delete(0, tk.END)
        dm_entry.delete(0, tk.END)
        daa_entry.delete(0, tk.END)
        os_entry.delete(0, tk.END)
        coa_entry.delete(0, tk.END)
        evs_entry.delete(0, tk.END)
        pdp_entry.delete(0, tk.END)
    except PermissionError:
        messagebox.showerror("Error", "The Excel file is currently open. Please close it and try again.")

def get_row_data(serial_no):
    filepath = r"C:\Users\infin\.jupyter\Book1.xlsx"
    
    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        
        # Iterate through rows to find the row with the matching serial number
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=9):
            if str(row[0].value) == str(serial_no):
                # Extract data from the row
                data = {
                    "Student Name": row[1].value,
                    "Enrollment ID": row[2].value,
                    "DM Score": row[3].value,
                    "DAA Score": row[4].value,
                    "OS Score": row[5].value,
                    "COA Score": row[6].value,
                    "EVS Score": row[7].value,
                    "PDP Score": row[8].value
                }
                return data
        
        # If serial number not found, return None
        return None
    except Exception as e:
        print("Error:", e)
        return None

def load_selected_row_for_edit():
    # Get the serial number from the entry
    serial_no = selected_serial_no_entry.get()

    if serial_no:  # Check if a serial number is provided
        # Get data of selected row using get_row_data function
        row_data = get_row_data(serial_no)
        
        if row_data:
            # Populate entry fields with data from the selected row
            name_entry_edit.delete(0, tk.END)
            name_entry_edit.insert(0, row_data["Student Name"])
            
            id_entry_edit.delete(0, tk.END)
            id_entry_edit.insert(0, row_data["Enrollment ID"])
            
            dm_entry_edit.delete(0, tk.END)
            dm_entry_edit.insert(0, row_data["DM Score"])
            
            daa_entry_edit.delete(0, tk.END)
            daa_entry_edit.insert(0, row_data["DAA Score"])
            
            os_entry_edit.delete(0, tk.END)
            os_entry_edit.insert(0, row_data["OS Score"])
            
            coa_entry_edit.delete(0, tk.END)
            coa_entry_edit.insert(0, row_data["COA Score"])
            
            evs_entry_edit.delete(0, tk.END)
            evs_entry_edit.insert(0, row_data["EVS Score"])
            
            pdp_entry_edit.delete(0, tk.END)
            pdp_entry_edit.insert(0, row_data["PDP Score"])
        else:
            messagebox.showerror("Error", f"No data found for serial number {serial_no}.")
    else:
        messagebox.showerror("Error", "Please enter a serial number.")



# Function to save changes made in update/edit section
def save_changes_after_edit():
    # Get edited data from entry fields
    edited_data = {
        "Student Name": name_entry_edit.get(),
        "Enrollment ID": id_entry_edit.get(),
        "DM Score": dm_entry_edit.get(),
        "DAA Score": daa_entry_edit.get(),
        "OS Score": os_entry_edit.get(),
        "COA Score": coa_entry_edit.get(),
        "EVS Score": evs_entry_edit.get(),
        "PDP Score": pdp_entry_edit.get()
    }

    if not all(edited_data.values()):
        messagebox.showerror("Error", "All fields must be filled.")
        return
    
    # Get serial number of selected row
    selected_serial_no = selected_serial_no_entry.get()

    # Update data of selected row in Excel file with edited data
    filepath = r"C:\Users\infin\.jupyter\Book1.xlsx"
    
    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        
        # Find the row corresponding to the selected serial number
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=9):
            if row[0].value == int(selected_serial_no):
                # Update the row with edited data
                row[1].value = edited_data["Student Name"]
                row[2].value = edited_data["Enrollment ID"]
                row[3].value = edited_data["DM Score"]
                row[4].value = edited_data["DAA Score"]
                row[5].value = edited_data["OS Score"]
                row[6].value = edited_data["COA Score"]
                row[7].value = edited_data["EVS Score"]
                row[8].value = edited_data["PDP Score"]
                break
        
        # Save the workbook
        wb.save(filepath)
        print("Changes saved successfully!")

        # Clear entry fields after saving changes
        name_entry_edit.delete(0, tk.END)
        id_entry_edit.delete(0, tk.END)
        dm_entry_edit.delete(0, tk.END)
        daa_entry_edit.delete(0, tk.END)
        os_entry_edit.delete(0, tk.END)
        coa_entry_edit.delete(0, tk.END)
        evs_entry_edit.delete(0, tk.END)
        pdp_entry_edit.delete(0, tk.END)
        
    except PermissionError:
        messagebox.showerror("Error", "The Excel file is currently open. Please close it and try again.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")



# Create main window
root = tk.Tk()
root.title("Student Data Entry")

# Insert Section
insert_frame = tk.LabelFrame(root, text="Insert Data")
insert_frame.grid(row=0, column=0, padx=10, pady=10)

# Create and place labels and entry fields for insert section
tk.Label(insert_frame, text="Student Name:").grid(row=0, column=0, padx=5, pady=5)
name_entry = tk.Entry(insert_frame)
name_entry.grid(row=0, column=1, padx=5, pady=5)

tk.Label(insert_frame, text="Enrollment ID:").grid(row=1, column=0, padx=5, pady=5)
id_entry = tk.Entry(insert_frame)
id_entry.grid(row=1, column=1, padx=5, pady=5)

tk.Label(insert_frame, text="DM Score:").grid(row=2, column=0, padx=5, pady=5)
dm_entry = tk.Entry(insert_frame)
dm_entry.grid(row=2, column=1, padx=5, pady=5)

tk.Label(insert_frame, text="DAA Score:").grid(row=3, column=0, padx=5, pady=5)
daa_entry = tk.Entry(insert_frame)
daa_entry.grid(row=3, column=1, padx=5, pady=5)

tk.Label(insert_frame, text="OS Score:").grid(row=4, column=0, padx=5, pady=5)
os_entry = tk.Entry(insert_frame)
os_entry.grid(row=4, column=1, padx=5, pady=5)

tk.Label(insert_frame, text="COA Score:").grid(row=5, column=0, padx=5, pady=5)
coa_entry = tk.Entry(insert_frame)
coa_entry.grid(row=5, column=1, padx=5, pady=5)

tk.Label(insert_frame, text="EVS Score:").grid(row=6, column=0, padx=5, pady=5)
evs_entry = tk.Entry(insert_frame)
evs_entry.grid(row=6, column=1, padx=5, pady=5)

tk.Label(insert_frame, text="PDP Score:").grid(row=7, column=0, padx=5, pady=5)
pdp_entry = tk.Entry(insert_frame)
pdp_entry.grid(row=7, column=1, padx=5, pady=5)

# Entry field for serial number of selected row
selected_serial_no_label = tk.Label(insert_frame, text="Serial No:")
selected_serial_no_label.grid(row=8, column=0, padx=5, pady=5)
selected_serial_no_entry = tk.Entry(insert_frame)
selected_serial_no_entry.grid(row=8, column=1, padx=5, pady=5)

# Button to save data for insertion
insert_button = tk.Button(insert_frame, text="Insert Data", command=save_data)
insert_button.grid(row=8, column=0, columnspan=2, pady=5)

# Update/Edit Section
edit_frame = tk.LabelFrame(root, text="Update/Edit Data")
edit_frame.grid(row=0, column=1, padx=10, pady=10)

# Entry field for serial number of selected row
selected_serial_no_label = tk.Label(edit_frame, text="Serial No:")
selected_serial_no_label.grid(row=0, column=0, padx=5, pady=5)
selected_serial_no_entry = tk.Entry(edit_frame)
selected_serial_no_entry.grid(row=0, column=1, padx=5, pady=5)

# Button to load selected row data into GUI for update/edit
load_button = tk.Button(edit_frame, text="Load Selected Row", command=load_selected_row_for_edit)
load_button.grid(row=1, column=0, padx=5, pady=5, columnspan=2)

# Entry fields for update/edit section
name_label = tk.Label(edit_frame, text="Student Name:")
name_label.grid(row=2, column=0, padx=5, pady=5)
name_entry_edit = tk.Entry(edit_frame)
name_entry_edit.grid(row=2, column=1, padx=5, pady=5)

id_label = tk.Label(edit_frame, text="Enrollment ID:")
id_label.grid(row=3, column=0, padx=5, pady=5)
id_entry_edit = tk.Entry(edit_frame)
id_entry_edit.grid(row=3, column=1, padx=5, pady=5)

dm_label = tk.Label(edit_frame, text="DM Score:")
dm_label.grid(row=4, column=0, padx=5, pady=5)
dm_entry_edit = tk.Entry(edit_frame)
dm_entry_edit.grid(row=4, column=1, padx=5, pady=5)

daa_label = tk.Label(edit_frame, text="DAA Score:")
daa_label.grid(row=5, column=0, padx=5, pady=5)
daa_entry_edit = tk.Entry(edit_frame)
daa_entry_edit.grid(row=5, column=1, padx=5, pady=5)

os_label = tk.Label(edit_frame, text="OS Score:")
os_label.grid(row=6, column=0, padx=5, pady=5)
os_entry_edit = tk.Entry(edit_frame)
os_entry_edit.grid(row=6, column=1, padx=5, pady=5)

coa_label = tk.Label(edit_frame, text="COA Score:")
coa_label.grid(row=7, column=0, padx=5, pady=5)
coa_entry_edit = tk.Entry(edit_frame)
coa_entry_edit.grid(row=7, column=1, padx=5, pady=5)

evs_label = tk.Label(edit_frame, text="EVS Score:")
evs_label.grid(row=8, column=0, padx=5, pady=5)
evs_entry_edit = tk.Entry(edit_frame)
evs_entry_edit.grid(row=8, column=1, padx=5, pady=5)

pdp_label = tk.Label(edit_frame, text="PDP Score:")
pdp_label.grid(row=9, column=0, padx=5, pady=5)
pdp_entry_edit = tk.Entry(edit_frame)
pdp_entry_edit.grid(row=9, column=1, padx=5, pady=5)


# Button to save changes made in update/edit section
save_button = tk.Button(edit_frame, text="Save Changes", command=save_changes_after_edit)
save_button.grid(row=11, column=0, columnspan=2, pady=5)

root.mainloop()

    